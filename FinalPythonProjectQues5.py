import openpyxl
import pandas as pd

excel_file = 'housing+(1).xlsx'
workbook = openpyxl.load_workbook(excel_file)
sheet = workbook['housing (2) (1)']

# Filling missing data with the mean value of total_bedrooms
column_names = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]  # Getting the column names
total_bedrooms_index = column_names.index('total_bedrooms')  # Getting the index of total_bedrooms
total_bedrooms_column = [row[total_bedrooms_index].value for row in sheet.iter_rows(min_row=2)]  # Getting the total_bedrooms values
mean_total_bedrooms = pd.Series(total_bedrooms_column).mean()  # Getting the mean value of total_bedrooms
for row in sheet.iter_rows(min_row=2):   # Iterating over the rows
    if row[total_bedrooms_index].value is None:  # If the total_bedrooms value is missing
        row[total_bedrooms_index].value = mean_total_bedrooms   # Filling it with the mean value
new_excel_file = "new_housingQues5.xlsx"  # Creating a new excel file
workbook.save(new_excel_file)  # Saving the new excel file
print(mean_total_bedrooms)   # Printing the mean value of total_bedrooms
